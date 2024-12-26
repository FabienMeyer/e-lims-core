"""Device under test generatre excel worksheet module."""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.styles import Alignment, Border, Font, Side

from e_lims_core.utils.dut.tray import Tray

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

THICK = Side(border_style='thick')
THIN = Side(border_style='thin')
HEADER = Border(top=THICK, right=THICK, bottom=THICK, left=THICK)
ROW_INDEX_LEFT = Border(top=THICK, right=THIN, bottom=THICK, left=THICK)
ROW_INDEX_RIGHT = Border(top=THICK, right=THICK, bottom=THICK, left=THIN)
ROW_INDEX_MIDDLE = Border(top=THICK, right=THIN, bottom=THICK, left=THIN)
COL_INDEX_TOP = Border(top=THICK, right=THICK, bottom=THIN, left=THICK)
COL_INDEX_BOTTOM = Border(top=THIN, right=THICK, bottom=THICK, left=THICK)
COL_INDEX_MIDDLE = Border(top=THIN, right=THICK, bottom=THIN, left=THICK)
SQUARE_CORNER_TOP_LEFT = Border(top=THICK, right=THIN, bottom=THIN, left=THICK)
SQUARE_CORNER_TOP_RIGHT = Border(top=THICK, right=THICK, bottom=THIN, left=THIN)
SQUARE_CORNER_BOTTOM_LEFT = Border(top=THIN, right=THIN, bottom=THICK, left=THICK)
SQUARE_CORNER_BOTTOM_RIGHT = Border(top=THIN, right=THICK, bottom=THICK, left=THIN)
SQUARE_TOP = Border(top=THICK, right=THIN, bottom=THIN, left=THIN)
SQUARE_BOTTOM = Border(top=THIN, right=THIN, bottom=THICK, left=THIN)
SQUARE_LEFT = Border(top=THIN, right=THIN, bottom=THIN, left=THICK)
SQUARE_RIGHT = Border(top=THIN, right=THICK, bottom=THIN, left=THIN)
SQUARE_MIDDLE = Border(top=THIN, right=THIN, bottom=THIN, left=THIN)


class Tray2Excel:
    """Represents a tray of devices under test (DUT) in an Excel file."""

    HEADER_START_ROW = 1
    HEADER_START_COL = 1
    COLUMN_START_ROW = HEADER_START_ROW + 1
    COLUMN_START_COL = HEADER_START_COL + 1
    ROW_START_ROW = COLUMN_START_ROW + 1
    ROW_START_COL = HEADER_START_COL
    DATA_ROW_START = ROW_START_ROW
    DATA_COL_START = COLUMN_START_COL

    def __init__(self, tray: Tray, workbook: Workbook) -> None:
        """Initialize the Tray2Excel object."""
        self.tray = tray
        self.workbook = workbook
        self.worksheet = self.creat_and_active_worksheet()

    def generate(self) -> Workbook:
        """Export the tray to an Excel file."""
        self.create_and_format_title()
        self.create_and_format_columns()
        self.create_and_format_rows()
        self.create_and_format_data()
        return self.workbook

    def creat_and_active_worksheet(self) -> Worksheet:
        """Create a worksheet for the tray."""
        self.workbook.create_sheet(title=self.tray.name)
        return self.workbook[self.tray.name]

    def get_title_format(self) -> tuple[Font, Alignment, Border]:
        """Get the title format for the tray."""
        font = Font(bold=True, color='00FF0000', size=20)
        alignment = Alignment(horizontal='center', vertical='center')
        return font, alignment, HEADER

    def create_and_format_title(self) -> None:
        """Create the title format for the tray."""
        title = self.tray.name.replace('_', ' ').capitalize()
        _, columns = self.tray.get_tray().shape
        cell = self.worksheet.cell(
            row=self.HEADER_START_ROW,
            column=self.HEADER_START_COL,
            value=title,
        )
        cell.font, cell.alignment, cell.border = self.get_title_format()
        self.worksheet.merge_cells(
            start_row=self.HEADER_START_ROW,
            start_column=self.HEADER_START_COL,
            end_row=self.HEADER_START_ROW,
            end_column=self.HEADER_START_COL + columns,
        )

    def get_columns_format(self, columns: int, col: int) -> tuple[Font, Alignment, Border]:
        """Get the column format for the tray.

        Args:
        ----
        columns (int): The number of columns in the tray.
        col (int): The column index.

        """
        font = Font(bold=True, color='00000000', size=10)
        alignment = Alignment(horizontal='center', vertical='center')

        if col == 0:
            border = ROW_INDEX_LEFT
        elif col == columns - 1:
            border = ROW_INDEX_RIGHT
        else:
            border = ROW_INDEX_MIDDLE
        return font, alignment, border

    def create_and_format_columns(self) -> None:
        """Create the column format for the tray."""
        tray_df = self.tray.get_tray()
        _, columns = tray_df.shape

        for col in range(columns):
            cell = self.worksheet.cell(
                row=self.COLUMN_START_ROW,
                column=self.COLUMN_START_COL + col,
                value=tray_df.columns[col],
            )
            cell.font, cell.alignment, cell.border = self.get_columns_format(columns, col)

    def get_rows_format(self, rows: int, row: int) -> tuple[Font, Alignment, Border]:
        """Get the row format for the tray.

        Args:
        ----
        rows (int): The number of rows in the tray.
        columns (int): The number of columns in the tray.
        row (int): The row index.
        col (int): The column index.

        """
        font = Font(bold=True, color='00000000', size=10)
        alignment = Alignment(horizontal='center', vertical='center')

        if row == 0:
            border = COL_INDEX_TOP
        elif row == rows - 1:
            border = COL_INDEX_BOTTOM
        else:
            border = COL_INDEX_MIDDLE
        return font, alignment, border

    def create_and_format_rows(self) -> None:
        """Create the row format for the tray."""
        tray_df = self.tray.get_tray()
        rows, _ = tray_df.shape
        for row in range(rows):
            cell = self.worksheet.cell(
                row=self.ROW_START_ROW + row,
                column=self.ROW_START_COL,
                value=tray_df.index[row],
            )
            cell.font, cell.alignment, cell.border = self.get_rows_format(rows, row)

    def get_data_format_data(self, rows: int, columns: int, row: int, col: int) -> tuple[Font, Alignment, Border]:
        """Get the data format for the tray.

        Args:
        ----
        rows (int): The number of rows in the tray.
        columns (int): The number of columns in the tray.
        row (int): The row index.
        col (int): The column index.

        """
        font = Font(bold=False, color='00000000', size=10)
        alignment = Alignment(horizontal='center', vertical='center')
        if row == 0 and col == 0:
            border = SQUARE_CORNER_TOP_LEFT
        elif row == 0 and col == columns - 1:
            border = SQUARE_CORNER_TOP_RIGHT
        elif row == rows - 1 and col == 0:
            border = SQUARE_CORNER_BOTTOM_LEFT
        elif row == rows - 1 and col == columns - 1:
            border = SQUARE_CORNER_BOTTOM_RIGHT
        elif row == 0 and col != columns - 1 and col != 0:
            border = SQUARE_TOP
        elif row == rows - 1 and col != columns - 1 and col != 0:
            border = SQUARE_BOTTOM
        elif col == 0 and row != rows - 1 and row != 0:
            border = SQUARE_LEFT
        elif col == columns - 1 and row != rows - 1 and row != 0:
            border = SQUARE_RIGHT
        else:
            border = SQUARE_MIDDLE
        return font, alignment, border

    def create_and_format_data(self) -> None:
        """Create the data format for the tray."""
        tray_df = self.tray.get_tray()
        rows, columns = tray_df.shape
        for col in range(columns):
            for row in range(rows):
                cell = self.worksheet.cell(
                    row=self.DATA_ROW_START + row,
                    column=self.DATA_COL_START + col,
                    value=str(tray_df.iloc[row, col]),
                )
                cell.font, cell.alignment, cell.border = self.get_data_format_data(rows, columns, row, col)

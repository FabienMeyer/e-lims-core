"""Module for testing the export to XLSX functionality."""

import math

from openpyxl import Workbook

from e_lims_core.utils.dut.export.tray2xlsx import (
    COL_INDEX_BOTTOM,
    COL_INDEX_MIDDLE,
    COL_INDEX_TOP,
    HEADER,
    ROW_INDEX_LEFT,
    ROW_INDEX_MIDDLE,
    ROW_INDEX_RIGHT,
    SQUARE_BOTTOM,
    SQUARE_CORNER_BOTTOM_LEFT,
    SQUARE_CORNER_BOTTOM_RIGHT,
    SQUARE_CORNER_TOP_LEFT,
    SQUARE_CORNER_TOP_RIGHT,
    SQUARE_LEFT,
    SQUARE_MIDDLE,
    SQUARE_RIGHT,
    SQUARE_TOP,
    Tray2Excel,
)
from e_lims_core.utils.dut.tray import Tray


def test_tray2excel_init(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the initialization of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    assert tray2excel.tray == fx_tray
    assert tray2excel.workbook == fx_workbook
    assert tray2excel.worksheet.title == fx_tray.name
    """Module for testing the export to XLSX functionality."""


def test_tray2excel_generate(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the generate method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    workbook = tray2excel.generate()
    worksheet = workbook[fx_tray.name]

    # Check title
    title_cell = worksheet.cell(row=tray2excel.HEADER_START_ROW, column=tray2excel.HEADER_START_COL)
    assert title_cell.value == fx_tray.name.replace('_', ' ').capitalize()

    # Check columns
    tray_df = fx_tray.get_tray()
    _, columns = tray_df.shape
    for col in range(columns):
        cell = worksheet.cell(row=tray2excel.COLUMN_START_ROW, column=tray2excel.COLUMN_START_COL + col)
        assert cell.value == tray_df.columns[col]

    # Check rows
    rows, _ = tray_df.shape
    for row in range(rows):
        cell = worksheet.cell(row=tray2excel.ROW_START_ROW + row, column=tray2excel.ROW_START_COL)
        assert cell.value == tray_df.index[row]

    # Check data
    for col in range(columns):
        for row in range(rows):
            cell = worksheet.cell(row=tray2excel.DATA_ROW_START + row, column=tray2excel.DATA_COL_START + col)
            assert cell.value == str(tray_df.iloc[row, col])


def test_creat_and_active_worksheet(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the creat_and_active_worksheet method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    worksheet = tray2excel.creat_and_active_worksheet()
    assert worksheet.title == fx_tray.name
    assert worksheet == fx_workbook[fx_tray.name]


def test_get_title_format(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the get_title_format method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    font, alignment, border = tray2excel.get_title_format()

    assert font.bold is True
    assert font.color is not None
    assert font.color.value == '00FF0000'
    assert math.isclose(font.size, 20.0, rel_tol=1e-9)

    assert alignment.horizontal == 'center'
    assert alignment.vertical == 'center'

    assert border == HEADER


def test_get_columns_format(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the get_columns_format method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    tray_df = fx_tray.get_tray()
    _, columns = tray_df.shape

    for col in range(columns):
        font, alignment, border = tray2excel.get_columns_format(columns, col)

        assert font.bold is True
        assert font.color is not None
        assert font.color.value == '00000000'
        assert math.isclose(font.size, 10.0, rel_tol=1e-9)

        assert alignment.horizontal == 'center'

        if col == 0:
            assert border == ROW_INDEX_LEFT
        elif col == columns - 1:
            assert border == ROW_INDEX_RIGHT
        else:
            assert border == ROW_INDEX_MIDDLE
        assert alignment.vertical == 'center'


def test_get_rows_format(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the get_rows_format method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    tray_df = fx_tray.get_tray()
    rows, _ = tray_df.shape

    for row in range(rows):
        font, alignment, border = tray2excel.get_rows_format(rows, row)

        assert font.bold is True
        assert font.color is not None
        assert font.color.value == '00000000'
        assert math.isclose(font.size, 10.0, rel_tol=1e-9)

        assert alignment.horizontal == 'center'
        assert alignment.vertical == 'center'

        if row == 0:
            assert border == COL_INDEX_TOP
        elif row == rows - 1:
            assert border == COL_INDEX_BOTTOM
        else:
            assert border == COL_INDEX_MIDDLE


def test_create_and_format_rows(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the create_and_format_rows method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    tray2excel.create_and_format_rows()
    tray_df = fx_tray.get_tray()
    rows, _ = tray_df.shape

    for row in range(rows):
        cell = tray2excel.worksheet.cell(
            row=tray2excel.ROW_START_ROW + row,
            column=tray2excel.ROW_START_COL,
        )
        assert cell.value == tray_df.index[row]
        font, alignment, border = tray2excel.get_rows_format(rows, row)
        assert cell.font == font
        assert cell.alignment == alignment
        assert cell.border == border


def test_get_data_format_data(fx_tray: Tray, fx_workbook: Workbook) -> None:  # noqa: C901
    """Test the get_data_format_data method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    tray_df = fx_tray.get_tray()
    rows, columns = tray_df.shape

    for row in range(rows):
        for col in range(columns):
            font, alignment, border = tray2excel.get_data_format_data(rows, columns, row, col)

            assert font.bold is False
            assert font.color is not None
            assert font.color.value == '00000000'
            assert math.isclose(font.size, 10.0, rel_tol=1e-9)

            assert alignment.horizontal == 'center'
            assert alignment.vertical == 'center'

            if row == 0 and col == 0:
                assert border == SQUARE_CORNER_TOP_LEFT
            elif row == 0 and col == columns - 1:
                assert border == SQUARE_CORNER_TOP_RIGHT
            elif row == rows - 1 and col == 0:
                assert border == SQUARE_CORNER_BOTTOM_LEFT
            elif row == rows - 1 and col == columns - 1:
                assert border == SQUARE_CORNER_BOTTOM_RIGHT
            elif row == 0 and col != columns - 1 and col != 0:
                assert border == SQUARE_TOP
            elif row == rows - 1 and col != columns - 1 and col != 0:
                assert border == SQUARE_BOTTOM
            elif col == 0 and row != rows - 1 and row != 0:
                assert border == SQUARE_LEFT
            elif col == columns - 1 and row != rows - 1 and row != 0:
                assert border == SQUARE_RIGHT
            else:
                assert border == SQUARE_MIDDLE


def test_create_and_format_data(fx_tray: Tray, fx_workbook: Workbook) -> None:
    """Test the create_and_format_data method of the Tray2Excel object.

    Args:
    ----
        fx_tray (Tray): Fixture for creating a tray.
        fx_workbook (Workbook): Fixture for creating a workbook.

    """
    tray2excel = Tray2Excel(fx_tray, fx_workbook)
    tray2excel.create_and_format_data()
    tray_df = fx_tray.get_tray()
    rows, columns = tray_df.shape

    for col in range(columns):
        for row in range(rows):
            cell = tray2excel.worksheet.cell(
                row=tray2excel.DATA_ROW_START + row,
                column=tray2excel.DATA_COL_START + col,
            )
            assert cell.value == str(tray_df.iloc[row, col])
            font, alignment, border = tray2excel.get_data_format_data(rows, columns, row, col)
            assert cell.font == font
            assert cell.alignment == alignment
            assert cell.border == border
